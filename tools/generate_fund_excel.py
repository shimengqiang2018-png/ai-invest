#!/usr/bin/env python3
"""天天基金持仓截图 → Excel 生成工具

配合 fund-screenshot-ocr skill 使用。
从多张截图中提取的结构化数据生成带格式的 Excel 文件。

funds.json 格式：
[
  {
    "code": "013308",
    "name": "易方达恒生科技ETF联接(QDII)A",
    "market_value": 6607.32,
    "total_return": 102.38,
    "return_rate": 0.0157,    // 小数，0.0157 = 1.57%
    "daily_return": 296.50,
    "source_img": "图1",       // 来源图片编号
    "note": ""                 // 备注（如"双次校验"）
  }
]
"""

import json
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_fund_excel(
    funds: list[dict],
    output_path: str,
    image_files: list[str] | None = None,
    image_meta: list[dict] | None = None,
    screenshot_date: str | None = None,
):
    """
    Parameters
    ----------
    funds : list[dict]
        每只基金一条记录，字段：
        - code: 基金代码
        - name: 基金名称
        - market_value: 持仓市值
        - total_return: 累计收益
        - return_rate: 持仓收益率（小数，如 0.0273 = 2.73%）
        - daily_return: 昨日收益（可选）
        - source_img: 来源图片编号（可选）
        - note: 备注（可选）
    output_path : str
        输出 Excel 文件路径
    image_files : list[str] | None
        原始截图文件名列表（向后兼容）
    image_meta : list[dict] | None
        每张图片的处理明细：
        - file: 文件名
        - status: "已读取" / "部分识别" / "识别失败"
        - fund_count: 从该图识别出的基金数量
        - note: 备注（OCR 质量问题等）
    screenshot_date : str | None
        截图日期 (YYYY-MM-DD)，默认今天
    """
    wb = openpyxl.Workbook()

    # ── 样式定义 ──────────────────────────────────
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    data_align = Alignment(horizontal="right", vertical="center")
    text_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    green_font = Font(color="006100")
    red_font = Font(color="CC0000")
    sum_font = Font(bold=True, size=11)

    # ── Sheet 1: 持仓明细 ─────────────────────────
    ws1 = wb.active
    ws1.title = "持仓明细"

    headers = [
        "基金代码", "基金名称", "持仓市值", "累计收益",
        "持仓收益率", "昨日收益", "来源图片", "备注",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = header_align

    for i, f in enumerate(funds, 2):
        # 基金代码
        c = ws1.cell(row=i, column=1, value=f.get("code", ""))
        c.border = thin_border
        c.alignment = text_align

        # 基金名称
        c = ws1.cell(row=i, column=2, value=f.get("name", ""))
        c.border = thin_border
        c.alignment = text_align

        # 持仓市值
        mv = f.get("market_value")
        c = ws1.cell(row=i, column=3, value=mv)
        c.border = thin_border
        c.alignment = data_align
        c.number_format = "#,##0.00"

        # 累计收益（红绿色）
        total_ret = f.get("total_return", 0) or 0
        c = ws1.cell(row=i, column=4, value=total_ret)
        c.border = thin_border
        c.alignment = data_align
        c.number_format = "#,##0.00"
        c.font = green_font if total_ret >= 0 else red_font

        # 持仓收益率（红绿色 + 百分比格式）
        ret_rate = f.get("return_rate")
        c = ws1.cell(row=i, column=5, value=ret_rate)
        c.border = thin_border
        c.alignment = data_align
        if ret_rate is not None:
            c.number_format = "0.00%"
            c.font = green_font if ret_rate >= 0 else red_font

        # 昨日收益
        daily = f.get("daily_return", 0) or 0
        c = ws1.cell(row=i, column=6, value=daily)
        c.border = thin_border
        c.alignment = data_align
        c.number_format = "#,##0.00"
        c.font = green_font if daily >= 0 else red_font

        # 来源图片
        c = ws1.cell(row=i, column=7, value=f.get("source_img", ""))
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")

        # 备注
        c = ws1.cell(row=i, column=8, value=f.get("note", ""))
        c.border = thin_border
        c.alignment = text_align

    # 合计行
    last_data_row = len(funds) + 1
    total_row = last_data_row + 1
    sum_cols = [3, 4, 6]  # 持仓市值、累计收益、昨日收益

    for col in sum_cols:
        col_letter = get_column_letter(col)
        c = ws1.cell(
            row=total_row, column=col,
            value=f"=SUM({col_letter}2:{col_letter}{last_data_row})",
        )
        c.font = sum_font
        c.border = thin_border
        c.alignment = data_align
        c.number_format = "#,##0.00"

    c = ws1.cell(row=total_row, column=2, value="合计")
    c.font = sum_font
    c.border = thin_border
    c.alignment = text_align

    # 列宽
    col_widths = [12, 32, 14, 14, 14, 14, 10, 20]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # 冻结首行 + 自动筛选
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{total_row}"

    # ── Sheet 2: 汇总统计 ─────────────────────────
    ws2 = wb.create_sheet("汇总统计")

    total_mv = sum(f.get("market_value", 0) or 0 for f in funds)
    total_return = sum(f.get("total_return", 0) or 0 for f in funds)
    total_cost = total_mv - total_return
    total_rate = total_return / total_cost if total_cost != 0 else 0
    total_daily = sum(f.get("daily_return", 0) or 0 for f in funds)

    summary_data = [
        ("总持仓市值", f"¥{total_mv:,.2f}"),
        ("总投入成本", f"¥{total_cost:,.2f}"),
        ("总累计收益", f"{'+¥' if total_return >= 0 else '-¥'}{abs(total_return):,.2f}"),
        ("总收益率", f"{total_rate:.2%}"),
        ("昨日总收益", f"{'+¥' if total_daily >= 0 else '-¥'}{abs(total_daily):,.2f}"),
        ("持仓基金数", f"{len(funds)} 只"),
        ("来源图片数", f"{len(image_files or image_meta or [])} 张"),
        ("数据日期", screenshot_date or date.today().strftime("%Y-%m-%d")),
    ]

    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)

    for i, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=i, column=1, value=label).font = label_font
        vc = ws2.cell(row=i, column=2, value=value)
        vc.font = value_font
        if label in ("总累计收益", "总收益率", "昨日总收益"):
            vc.font = green_font if (
                total_return >= 0 if label == "总累计收益"
                else total_rate >= 0 if label == "总收益率"
                else total_daily >= 0
            ) else red_font

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 22

    # ── Sheet 3: 图片处理记录 ─────────────────────
    ws3 = wb.create_sheet("图片处理记录")

    img_headers = ["文件名", "状态", "识别基金数", "备注"]
    for col, h in enumerate(img_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border

    if image_meta:
        for i, meta in enumerate(image_meta, 2):
            ws3.cell(row=i, column=1, value=meta.get("file", "")).border = thin_border
            ws3.cell(row=i, column=2, value=meta.get("status", "已读取")).border = thin_border
            ws3.cell(row=i, column=3, value=meta.get("fund_count", 0)).border = thin_border
            ws3.cell(row=i, column=4, value=meta.get("note", "")).border = thin_border
    elif image_files:
        for i, img in enumerate(image_files, 2):
            ws3.cell(row=i, column=1, value=img).border = thin_border
            ws3.cell(row=i, column=2, value="已读取").border = thin_border
            ws3.cell(row=i, column=3, value="N/A").border = thin_border
            ws3.cell(row=i, column=4, value="").border = thin_border

    ws3.column_dimensions["A"].width = 50
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 30

    # ── 保存 ─────────────────────────────────────
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel 已生成: {output_path}")
    print(f"   持仓基金: {len(funds)} 只")
    print(f"   总市值: ¥{total_mv:,.2f}")
    print(f"   总收益: {'+' if total_return >= 0 else '-'}¥{abs(total_return):,.2f}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python generate_fund_excel.py <funds.json> [output.xlsx]")
        print("")
        print("funds.json 格式:")
        print('[{"code":"013308","name":"易方达恒生科技ETF联接(QDII)A",')
        print('  "market_value":6607.32,"total_return":102.38,"return_rate":0.0157,')
        print('  "daily_return":296.50,"source_img":"图1","note":""}]')
        sys.exit(1)

    json_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else json_path.replace(".json", ".xlsx")

    with open(json_path) as f:
        funds = json.load(f)

    generate_fund_excel(funds, output_path, image_files=[json_path])
